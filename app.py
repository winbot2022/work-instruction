
import io
import os
import json
import base64
from datetime import date
from pathlib import Path

import matplotlib.pyplot as plt
from matplotlib import font_manager, rcParams
import pandas as pd
import streamlit as st
from openai import OpenAI

st.set_page_config(page_title="資金繰りMVP v11", layout="wide")

st.markdown("""
<style>
.small-note {font-size: 0.9rem; color: #666;}
.kpi-card {
    padding: 14px 18px;
    border-radius: 14px;
    background: #f7f9fc;
    border: 1px solid #d9e2f1;
    margin-bottom: 8px;
}
.kpi-title {
    font-size: 0.9rem;
    color: #555;
    margin-bottom: 6px;
}
.kpi-value {
    font-size: 1.8rem;
    font-weight: 700;
    line-height: 1.2;
}
</style>
""", unsafe_allow_html=True)

def setup_japanese_font():
    candidates = [
        Path(__file__).with_name("NotoSansJP-Regular.ttf"),
        Path.cwd() / "NotoSansJP-Regular.ttf",
    ]
    for font_path in candidates:
        if font_path.exists():
            try:
                font_manager.fontManager.addfont(str(font_path))
                font_name = font_manager.FontProperties(fname=str(font_path)).get_name()
                rcParams["font.family"] = font_name
                rcParams["axes.unicode_minus"] = False
                return ""
            except Exception:
                pass
    return None

font_msg = setup_japanese_font()

REQUIRED_COLS = ["date", "amount"]
SCHEDULE_REQUIRED_COLS = ["日付", "種別", "金額", "繰返し", "メモ"]

st.title("資金繰りMVP v11")
st.caption("銀行明細・通帳スキャン・予定Excelに対応。実績と予定を組み合わせて、向こう12か月の資金繰りを可視化します。")

def fmt(x):
    try:
        return f"{x:,.0f}"
    except Exception:
        return x

def fmt_man(x):
    try:
        return f"{x/10000:,.0f}"
    except Exception:
        return x

def read_csv_auto(file):
    file.seek(0)
    try:
        return pd.read_csv(file, encoding="utf-8-sig", header=None)
    except:
        file.seek(0)
        try:
            return pd.read_csv(file, encoding="cp932", header=None)
        except:
            file.seek(0)
            return pd.read_csv(file, header=None)

def read_csv_auto_with_header(file):
    try:
        return pd.read_csv(file, encoding="utf-8-sig")
    except Exception:
        file.seek(0)
        try:
            return pd.read_csv(file, encoding="cp932")
        except Exception:
            file.seek(0)
            return pd.read_csv(file, encoding="shift_jis")

def detect_bank_format(df: pd.DataFrame, uploaded_name: str = "") -> str:
    cols = [str(c).strip().lower() for c in df.columns]

    if {"date", "amount"}.issubset(set(cols)):
        return "standard"

    if {"日付", "金額"}.issubset(set(df.columns)):
        return "standard_jp"

    if df.shape[1] >= 1:
        detail_flag = pd.to_numeric(df.iloc[:, 0], errors="coerce")
        detail_df = df[detail_flag == 2].copy()

        if not detail_df.empty:
            if detail_df.shape[1] == 7:
                return "mufg"
            if detail_df.shape[1] == 6:
                return "shinkin_ib"

    return "unknown"

def parse_mufg_csv(df: pd.DataFrame) -> pd.DataFrame:
    x = df.copy()
    detail_flag = pd.to_numeric(x.iloc[:, 0], errors="coerce")
    x = x[detail_flag == 2].copy()

    x["date"] = pd.to_datetime(x.iloc[:, 1], errors="coerce")
    x["withdraw"] = pd.to_numeric(x.iloc[:, 4], errors="coerce").fillna(0)
    x["deposit"] = pd.to_numeric(x.iloc[:, 5], errors="coerce").fillna(0)
    x["amount"] = x["deposit"] - x["withdraw"]
    x["balance"] = pd.to_numeric(x.iloc[:, 6], errors="coerce")
    x["description"] = (
        x.iloc[:, 2].fillna("").astype(str).str.strip() + " " +
        x.iloc[:, 3].fillna("").astype(str).str.strip()
    ).str.strip()

    x = x.dropna(subset=["date"]).copy()
    x = x[(x["deposit"] != 0) | (x["withdraw"] != 0) | (x["balance"].notna())].copy()

    return x[["date", "amount", "balance", "description"]].sort_values("date").reset_index(drop=True)

def parse_shinkin_ib_csv(df: pd.DataFrame) -> pd.DataFrame:
    x = df.copy()
    detail_flag = pd.to_numeric(x.iloc[:, 0], errors="coerce")
    x = x[detail_flag == 2].copy()

    x["date"] = pd.to_datetime(x.iloc[:, 1], errors="coerce")
    x["description"] = x.iloc[:, 2].fillna("").astype(str).str.strip()
    x["withdraw"] = pd.to_numeric(x.iloc[:, 3], errors="coerce").fillna(0)
    x["deposit"] = pd.to_numeric(x.iloc[:, 4], errors="coerce").fillna(0)
    x["amount"] = x["deposit"] - x["withdraw"]
    x["balance"] = pd.to_numeric(x.iloc[:, 5], errors="coerce")

    x = x.dropna(subset=["date"]).copy()
    x = x[(x["deposit"] != 0) | (x["withdraw"] != 0) | (x["balance"].notna())].copy()

    return x[["date", "amount", "balance", "description"]].sort_values("date").reset_index(drop=True)



def normalize_standard_bank_df(df: pd.DataFrame) -> pd.DataFrame:
    rename_map = {}
    for src, dst in [("日付", "date"), ("金額", "amount"), ("摘要", "description"), ("残高", "balance")]:
        if src in df.columns:
            rename_map[src] = dst
    x = df.rename(columns=rename_map).copy()
    x.columns = [str(c).strip().lower() for c in x.columns]
    missing = [c for c in REQUIRED_COLS if c not in x.columns]
    if missing:
        raise ValueError(f"必要列が不足しています: {missing}")
    x["date"] = pd.to_datetime(x["date"], errors="coerce")
    x["amount"] = pd.to_numeric(x["amount"], errors="coerce")
    if "description" not in x.columns:
        x["description"] = ""
    if "balance" in x.columns:
        x["balance"] = pd.to_numeric(x["balance"], errors="coerce")
    return x.dropna(subset=["date", "amount"]).sort_values("date").reset_index(drop=True)

def parse_uploaded_bank_csv(uploaded_files):
    raw = read_csv_auto(uploaded_files)
    fmt_name = detect_bank_format(raw, getattr(uploaded_files, "name", ""))

    if fmt_name == "mufg":
        return parse_mufg_csv(raw), "MUFG形式として取り込みました。"

    if fmt_name == "shinkin_ib":
        return parse_shinkin_ib_csv(raw), "信金IB形式として取り込みました。"

    uploaded_files.seek(0)
    raw2 = read_csv_auto_with_header(uploaded_files)
    fmt_name = detect_bank_format(raw2, getattr(uploaded_files, "name", ""))

    if fmt_name in ("standard", "standard_jp"):
        return normalize_standard_bank_df(raw2), "標準CSV形式として取り込みました。"

    raise ValueError("CSV形式を判定できませんでした。今回は MUFG、信金IB、または date/amount 形式に対応しています。")

def image_bytes_to_data_url(image_bytes: bytes, file_name: str) -> str:
    ext = Path(file_name).suffix.lower()
    mime = "image/jpeg" if ext in [".jpg", ".jpeg"] else "image/png"
    b64 = base64.b64encode(image_bytes).decode("utf-8")
    return f"data:{mime};base64,{b64}"

def scan_passbook_to_json(image_bytes: bytes, file_name: str, model_name: str = "gpt-4.1"):
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY が設定されていません。PowerShellの環境変数を確認してください。")
    client = OpenAI(api_key=api_key)
    data_url = image_bytes_to_data_url(image_bytes, file_name)
    prompt = """
この通帳画像を読み取り、明細行だけを抽出してください。

各行について以下の項目を返してください。
- date: 日付
- withdrawal: 出金額（なければ 0）
- deposit: 入金額（なければ 0）
- balance: 残高
- description: 摘要

必ずJSON配列のみで返してください。
説明文やコードブロックは不要です。
数字のカンマは除去してください。
読めない項目は空欄または 0 としてください。
明細行以外（見出し、口座情報、注意書き）は除外してください。
date は画像記載のままで可です。
"""
    response = client.responses.create(
        model=model_name,
        input=[
            {
                "role": "user",
                "content": [
                    {"type": "input_text", "text": prompt},
                    {"type": "input_image", "image_url": data_url},
                ],
            }
        ],
    )
    return json.loads(response.output_text)

def normalize_passbook_json(rows: list, year_prefix: int = 2025) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=["date", "amount", "balance", "description"])
    df = pd.DataFrame(rows).copy()
    for col in ["withdrawal", "deposit", "balance"]:
        if col not in df.columns:
            df[col] = 0
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    if "description" not in df.columns:
        df["description"] = ""
    df["description"] = df["description"].fillna("").astype(str)
    if "date" not in df.columns:
        df["date"] = ""
    df["date_raw"] = df["date"].fillna("").astype(str).str.strip()

    def fix_date(s: str):
        s = str(s).strip().replace(".", "-").replace("/", "-")
        if not s:
            return pd.NaT
        if "-" in s:
            parts = [p for p in s.split("-") if p != ""]
            if len(parts) == 3:
                try:
                    y, m, d = [int(p) for p in parts]
                    year = 2018 + y if 1 <= y <= 30 else y
                    return pd.Timestamp(year=year, month=m, day=d)
                except Exception:
                    return pd.NaT
            elif len(parts) == 2:
                try:
                    y = int(parts[0])
                    md = parts[1]
                    year = 2018 + y if 1 <= y <= 30 else y
                    if len(md) == 3:
                        m = int(md[0])
                        d = int(md[1:])
                    elif len(md) == 4:
                        m = int(md[:2])
                        d = int(md[2:])
                    else:
                        return pd.NaT
                    return pd.Timestamp(year=year, month=m, day=d)
                except Exception:
                    return pd.NaT
        digits = "".join(ch for ch in s if ch.isdigit())
        if len(digits) == 6:
            try:
                y = int(digits[:2]); m = int(digits[2:4]); d = int(digits[4:6])
                year = 2018 + y if 1 <= y <= 30 else y
                return pd.Timestamp(year=year, month=m, day=d)
            except Exception:
                return pd.NaT
        return pd.NaT

    df["date"] = df["date_raw"].apply(fix_date)
    df["amount"] = df["deposit"] - df["withdrawal"]
    df = df[
        (df["date"].notna()) &
        (
            (df["deposit"] != 0) |
            (df["withdrawal"] != 0) |
            (df["balance"] != 0) |
            (df["description"].str.strip() != "")
        )
    ].copy()
    return df[["date", "amount", "balance", "description"]].sort_values("date").reset_index(drop=True)

def load_schedule_excel(file) -> pd.DataFrame:
    df = pd.read_excel(file)
    missing = [c for c in SCHEDULE_REQUIRED_COLS if c not in df.columns]
    if missing:
        raise ValueError(f"予定Excelの必要列が不足しています: {missing}")
    x = df.copy()
    x["date"] = pd.to_datetime(x["日付"], errors="coerce")
    x["kind"] = x["種別"].fillna("").astype(str).str.strip()
    x["amount"] = pd.to_numeric(x["金額"], errors="coerce")
    x["repeat"] = x["繰返し"].fillna("単発").astype(str).str.strip()
    x["memo"] = x["メモ"].fillna("").astype(str).str.strip()
    x = x[x["kind"].isin(["入金", "出金"])].copy()
    x = x[x["repeat"].isin(["単発", "毎月"])].copy()
    x = x.dropna(subset=["date", "amount"]).copy()
    return x[["date", "kind", "amount", "repeat", "memo"]].reset_index(drop=True)

def create_schedule_template_bytes():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "予定入力"

    # 上部説明
    ws["A1"] = "使い方"
    ws["A2"] = "1. 日付を入力  2. 種別と繰返しはプルダウンから選択  3. 金額を入力  4. メモを記入"
    ws["A3"] = "※ 標準の .xlsx では、カレンダー/電卓UIは環境依存です。テンプレではプルダウンと入力制約を設定しています。"

    # ▼ メモ候補シート追加（v11）
    master = wb.create_sheet("メモ候補")

    # 入金候補（最小構成）
    master["A1"] = "入金候補"
    master["A2"] = "売上"

    # 出金候補
    master["B1"] = "出金候補"

    out_list = [
        "仕入支払",
        "外注費",
        "人件費",
        "家賃",
        "設備投資",
        "借入返済",
        "利息支払",
        "税金",
        "賞与",
        "修繕費"
    ]

    for i, v in enumerate(out_list, start=2):
        master[f"B{i}"] = v

    ws["A1"].font = Font(bold=True)
    ws["A1"].fill = PatternFill("solid", fgColor="E2F0D9")
    ws["A2"].font = Font(color="666666")
    ws["A3"].font = Font(color="666666")

    # ヘッダー
    headers = ["日付", "種別", "金額", "繰返し", "メモ"]
    header_row = 4
    for col, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=h)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)

    # サンプル行
    sample_rows = [
        ["2026-05-01", "出金", 300000, "単発", "設備投資"],
        ["2026-05-25", "入金", 500000, "毎月", "売上"],
        ["2026-06-10", "出金", 120000, "単発", "賞与"],
    ]

    start_row = 5
    for r_idx, row in enumerate(sample_rows, start=start_row):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # 列幅
    widths = {
        "A": 14,  # 日付
        "B": 10,  # 種別
        "C": 14,  # 金額
        "D": 10,  # 繰返し
        "E": 28,  # メモ
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # 書式
    end_row = 203
    for row in range(start_row, end_row + 1):
        ws[f"A{row}"].number_format = "yyyy-mm-dd"
        ws[f"C{row}"].number_format = '#,##0'

    # プルダウン（直接指定の方が安定）
    dv_kind = DataValidation(type="list", formula1='"入金,出金"', allow_blank=True)
    dv_repeat = DataValidation(type="list", formula1='"単発,毎月"', allow_blank=True)

    ws.add_data_validation(dv_kind)
    ws.add_data_validation(dv_repeat)

    dv_kind.add(f"B{start_row}:B{end_row}")
    dv_repeat.add(f"D{start_row}:D{end_row}")

    # 日付入力制約
    date_dv = DataValidation(
        type="date",
        operator="between",
        formula1="DATE(2024,1,1)",
        formula2="DATE(2035,12,31)",
        allow_blank=True
    )
    date_dv.promptTitle = "日付入力"
    date_dv.prompt = "日付は yyyy-mm-dd 形式で入力してください。Excel環境によっては日付選択UIが出る場合があります。"
    ws.add_data_validation(date_dv)
    date_dv.add(f"A{start_row}:A{end_row}")

    # 金額入力制約
    amount_dv = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=True
    )
    amount_dv.promptTitle = "金額入力"
    amount_dv.prompt = "金額は円単位で入力してください。"
    ws.add_data_validation(amount_dv)
    amount_dv.add(f"C{start_row}:C{end_row}")

    # 固定表示
    ws.freeze_panes = "A5"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def monthly_base_flow(bank_df: pd.DataFrame, months_back: int = 6):
    latest = bank_df["date"].max()
    start = (latest - pd.DateOffset(months=months_back - 1)).replace(day=1)

    df = bank_df[bank_df["date"] >= start].copy()
    df["month"] = df["date"].dt.to_period("M").dt.to_timestamp()

    # 取引単位で入金・出金を分解
    df["inflow"] = df["amount"].clip(lower=0)
    df["outflow"] = (-df["amount"].clip(upper=0))

    # 月ごとに総入金・総出金を集計
    monthly = df.groupby("month", as_index=False)[["inflow", "outflow"]].sum()

    # 千円単位に丸め
    avg_inflow = round(monthly["inflow"].mean() / 1000) * 1000
    avg_outflow = round(monthly["outflow"].mean() / 1000) * 1000

    return float(avg_inflow), float(avg_outflow)

def build_event_df(df: pd.DataFrame):
    if df is None or df.empty:
        return pd.DataFrame(columns=["id", "date", "kind", "amount", "repeat", "memo"])

    x = df.copy()

    if "id" in x.columns:
        x["id"] = pd.to_numeric(x["id"], errors="coerce")

    x["date"] = pd.to_datetime(x["date"], errors="coerce")
    x["amount"] = pd.to_numeric(x["amount"], errors="coerce")

    x = x.dropna(subset=["id", "date", "amount"]).copy()
    x["id"] = x["id"].astype(int)

    return x[["id", "date", "kind", "amount", "repeat", "memo"]]

def expand_events(events_df: pd.DataFrame, horizon_months: int = 12):
    if events_df.empty:
        return (
            pd.DataFrame(columns=["month", "event_inflow", "event_outflow"]),
            pd.DataFrame(columns=["month", "id", "kind", "amount", "repeat", "memo"]),
        )

    rows, detail_rows = [], []
    horizon_start = pd.Timestamp(date.today().replace(day=1))
    horizon_end = horizon_start + pd.DateOffset(months=horizon_months - 1)

    for _, r in events_df.iterrows():
        current = pd.Timestamp(r["date"]).replace(day=1)

        while current <= horizon_end:
            if current >= horizon_start:

                inflow = 0.0
                outflow = 0.0

                if r["kind"] == "入金":
                    inflow = float(r["amount"])
                else:
                    outflow = float(r["amount"])

                rows.append({
                    "month": current,
                    "event_inflow": inflow,
                    "event_outflow": outflow
                })

                detail_rows.append({
                    "month": current,
                    "id": r["id"],
                    "kind": r["kind"],
                    "amount": float(r["amount"]),
                    "repeat": r["repeat"],
                    "memo": r["memo"],
                })

            if r["repeat"] != "毎月":
                break

            current = current + pd.DateOffset(months=1)

    monthly = (
        pd.DataFrame(rows)
        .groupby("month", as_index=False)[["event_inflow", "event_outflow"]]
        .sum()
        if rows else pd.DataFrame(columns=["month", "event_inflow", "event_outflow"])
    )

    detail = (
        pd.DataFrame(detail_rows)
        if detail_rows else pd.DataFrame(columns=["month", "id", "kind", "amount", "repeat", "memo"])
    )

    return monthly, detail

def build_forecast(bank_df: pd.DataFrame, events_df: pd.DataFrame, months_back=6, horizon_months=12, safety_threshold=0):
    avg_inflow, avg_outflow = monthly_base_flow(bank_df, months_back=months_back)
    if "balance" in bank_df.columns and bank_df["balance"].notna().any():
        if "source_file" in bank_df.columns:
            latest_balance_df = (
                bank_df.dropna(subset=["balance"])
                .sort_values("date")
                .groupby("source_file", as_index=False)
                .tail(1)
            )
            current_balance = float(latest_balance_df["balance"].sum())
        else:
            current_balance = float(bank_df["balance"].dropna().iloc[-1])
    else:
        current_balance = float(bank_df["amount"].sum())
    months = pd.date_range(pd.Timestamp(date.today().replace(day=1)), periods=horizon_months, freq="MS")
    forecast = pd.DataFrame({"month": months})
    forecast["base_inflow"] = avg_inflow
    forecast["base_outflow"] = avg_outflow
    event_monthly, event_detail = expand_events(events_df, horizon_months=horizon_months)
    forecast = forecast.merge(event_monthly, on="month", how="left")
    forecast[["event_inflow", "event_outflow"]] = forecast[["event_inflow", "event_outflow"]].fillna(0.0)
    forecast["inflow_total"] = forecast["base_inflow"] + forecast["event_inflow"]
    forecast["outflow_total"] = forecast["base_outflow"] + forecast["event_outflow"]
    forecast["net"] = forecast["inflow_total"] - forecast["outflow_total"]
    carry = current_balance
    carry_in, ending = [], []
    for _, r in forecast.iterrows():
        carry_in.append(carry)
        carry = carry + r["net"]
        ending.append(carry)
    forecast["carry_in"] = carry_in
    forecast["ending_balance"] = ending
    forecast["danger"] = forecast["ending_balance"] < safety_threshold
    return forecast, current_balance, event_detail

def event_summary_text(kind: str, amount_yen: int, repeat: str, memo: str):
    sign = "+" if kind == "入金" else "-"
    return f"{sign}{fmt(amount_yen)} 円 / {repeat} / {memo if memo else 'メモなし'}"

def format_event_label(r):
    return f"No.{r['id']} | {pd.Timestamp(r['date']).strftime('%Y-%m-%d')} | {r['kind']} | {fmt(r['amount'])}円 | {r['repeat']} | {r['memo'] if r['memo'] else 'メモなし'}"

def get_scale_info(max_yen):
    if max_yen >= 100000000:
        return 100000000, "億円"
    elif max_yen >= 1000000:
        return 1000000, "百万円"
    else:
        return 10000, "万円"

def analyze_danger_reason(forecast: pd.DataFrame, event_detail: pd.DataFrame):
    danger_df = forecast[forecast["danger"]]
    if danger_df.empty:
        return None, None
    first_danger = danger_df.iloc[0]
    target_month = first_danger["month"]
    reasons = [f"{target_month:%Y-%m} は通常収支が {fmt_man(first_danger['base_inflow'] - first_danger['base_outflow'])} 万円、イベント差分が {fmt_man(first_danger['event_inflow'] - first_danger['event_outflow'])} 万円です。"]
    month_events = event_detail[event_detail["month"] == target_month].copy() if not event_detail.empty else pd.DataFrame()
    cause_table = None
    if not month_events.empty:
        month_events["impact"] = month_events.apply(lambda r: r["amount"] if r["kind"] == "出金" else -r["amount"], axis=1)
        month_events = month_events.sort_values("impact", ascending=False)
        top_causes = month_events.head(3)
        cause_lines = []
        for _, r in top_causes.iterrows():
            label = r["memo"] if r["memo"] else "メモなし"
            cause_lines.append(f"{r['kind']}予定 No.{int(r['id'])}「{label}」が {fmt_man(r['amount'])} 万円")
        reasons.append("主な影響要因: " + "、".join(cause_lines))
        cause_table = top_causes[["id", "kind", "memo", "amount", "repeat"]].copy()
        cause_table["amount"] = cause_table["amount"].map(fmt)
        cause_table = cause_table.rename(columns={"id":"No.","kind":"種別","memo":"メモ","amount":"金額（円）","repeat":"繰返し"})
    else:
        reasons.append("この月に登録された個別イベントは見当たりません。通常収支の不足が主因です。")
    return first_danger, {"summary":"\n".join(reasons), "table":cause_table}

def build_project_forecast_from_events(initial_balance, events_df, horizon_months=12, safety_threshold=0):
    months = pd.date_range(
        pd.Timestamp(date.today().replace(day=1)),
        periods=horizon_months,
        freq="MS"
    )

    forecast = pd.DataFrame({"month": months})
    forecast["base_inflow"] = 0.0
    forecast["base_outflow"] = 0.0

    event_monthly, event_detail = expand_events(events_df, horizon_months=horizon_months)

    forecast = forecast.merge(event_monthly, on="month", how="left")
    forecast[["event_inflow", "event_outflow"]] = forecast[["event_inflow", "event_outflow"]].fillna(0.0)

    forecast["inflow_total"] = forecast["event_inflow"]
    forecast["outflow_total"] = forecast["event_outflow"]
    forecast["net"] = forecast["inflow_total"] - forecast["outflow_total"]

    carry = float(initial_balance)
    carry_in, ending = [], []

    for _, r in forecast.iterrows():
        carry_in.append(carry)
        carry = carry + r["net"]
        ending.append(carry)

    forecast["carry_in"] = carry_in
    forecast["ending_balance"] = ending
    forecast["danger"] = forecast["ending_balance"] < safety_threshold

    return forecast, float(initial_balance), event_detail

def highlight_cashflow(val):
    try:
        v = float(val)
        if v < 0:
            return "background-color: #fdeaea; color: #b00020; font-weight: 700;"
    except Exception:
        pass
    return ""

def highlight_row(row):
    styles = [""] * len(row)
    try:
        if row.name == "月末残高":
            for i, val in enumerate(row):
                v = float(val)
                if v < 0:
                    styles[i] = "background-color: #fdeaea; color: #b00020; font-weight: 700;"
    except Exception:
        pass
    return styles

if "amount_digits" not in st.session_state:
    st.session_state["amount_digits"] = ""
if "events_store" not in st.session_state or st.session_state["events_store"] is None:
    st.session_state["events_store"] = pd.DataFrame(columns=["id","date","kind","amount","repeat","memo"])
if "next_event_id" not in st.session_state:
    st.session_state["next_event_id"] = 1
if "edit_target_id" not in st.session_state:
    st.session_state["edit_target_id"] = None
if "bank_df_store" not in st.session_state:
    st.session_state["bank_df_store"] = None
if "scan_json_store" not in st.session_state:
    st.session_state["scan_json_store"] = None

def append_digits(s: str):
    st.session_state["amount_digits"] += s
def clear_digits():
    st.session_state["amount_digits"] = ""
def backspace_digits():
    st.session_state["amount_digits"] = st.session_state["amount_digits"][:-1]

with st.sidebar:
    st.header("設定")
    months_back = st.slider("ベース計算に使う過去月数", 3, 12, 6)
    horizon_months = st.slider("予測月数", 6, 12, 12)
    safety_threshold_man = st.number_input("安全残高（万円）", min_value=0, value=0, step=10)
    safety_threshold = safety_threshold_man * 10000
    scan_year_prefix = st.number_input("通帳スキャンの補完年", min_value=2000, max_value=2100, value=2025, step=1)
    scan_model = st.selectbox("通帳スキャンモデル", ["gpt-4.1", "gpt-4o"], index=0)

st.subheader("1. 実績データの取込")
tab1, tab2 = st.tabs(["銀行CSV", "通帳スキャン JPEG/PNG"])
bank_df = None

with tab1:
    
    uploaded_files = st.file_uploader(
        "銀行明細CSVをアップロードしてください",
        type=["csv"],
        accept_multiple_files=True,
        key="csv_upload"
    )

    # ▼▼▼ ここに追加 ▼▼▼
    st.markdown("### 前回データの読込")

    uploaded_project = st.file_uploader(
        "前回保存したExcelをアップロードしてください",
        type=["xlsx"],
        key="project_upload"
    )

    restore_project = st.button("この前回ファイルを復元", key="restore_project_btn")

    if uploaded_project is not None and restore_project:
        try:
            xls = pd.ExcelFile(uploaded_project)

            # 実績データ復元
            if "実績データ" in xls.sheet_names:
                restored_bank = pd.read_excel(xls, sheet_name="実績データ")

                if "date" in restored_bank.columns:
                    restored_bank["date"] = pd.to_datetime(restored_bank["date"], errors="coerce")
                if "amount" in restored_bank.columns:
                    restored_bank["amount"] = pd.to_numeric(restored_bank["amount"], errors="coerce")
                if "balance" in restored_bank.columns:
                    restored_bank["balance"] = pd.to_numeric(restored_bank["balance"], errors="coerce")

                st.session_state["bank_df_store"] = restored_bank
                st.success("実績データを復元しました")

            # 予定イベント復元
            if "予定イベント" in xls.sheet_names:
                restored_events = pd.read_excel(xls, sheet_name="予定イベント")

                if "date" in restored_events.columns:
                    restored_events["date"] = pd.to_datetime(restored_events["date"], errors="coerce")
                if "amount" in restored_events.columns:
                    restored_events["amount"] = pd.to_numeric(restored_events["amount"], errors="coerce")
                if "id" in restored_events.columns:
                    restored_events["id"] = pd.to_numeric(restored_events["id"], errors="coerce")

                expected_cols = ["id", "date", "kind", "amount", "repeat", "memo"]
                for col in expected_cols:
                    if col not in restored_events.columns:
                        restored_events[col] = None

                restored_events = restored_events[expected_cols].copy()
                restored_events = restored_events.dropna(subset=["id"]).copy()
                restored_events["id"] = restored_events["id"].astype(int)

                st.session_state["events_store"] = restored_events
                st.session_state["edit_target_id"] = None

                if not restored_events.empty:
                    st.session_state["next_event_id"] = int(restored_events["id"].max()) + 1
                else:
                    st.session_state["next_event_id"] = 1

                st.success("予定イベントを復元しました")

        except Exception as e:
            st.error(f"読込エラー: {e}")
    # ▲▲▲ ここまで ▲▲▲

    st.markdown('<div class="small-note">v11では銀行/信金CSV形式に対応しています。</div>', unsafe_allow_html=True)

    if uploaded_files:
        try:
            df_list = []
            messages = []

            for f in uploaded_files:
                df, parse_message = parse_uploaded_bank_csv(f)
                df["source_file"] = f.name
                df_list.append(df)
                messages.append(f"{f.name}: {parse_message}")

            bank_df_csv = pd.concat(df_list, ignore_index=True)
            bank_df_csv = bank_df_csv.sort_values("date").reset_index(drop=True)

            st.session_state["bank_df_store"] = bank_df_csv

            st.success(f"{len(bank_df_csv)} 件の銀行データを取り込みました。")

            for msg in messages:
                st.info(msg)

            with st.expander("取込データを確認"):
                preview = bank_df_csv.copy()
                preview["date"] = preview["date"].dt.strftime("%Y-%m-%d")
                if "balance" in preview.columns:
                    preview["balance"] = preview["balance"].map(fmt)
                preview["amount"] = preview["amount"].map(fmt)
                st.dataframe(preview.tail(50), use_container_width=True)

        except Exception as e:
            st.error(f"取込エラー: {e}")
    

with tab2:
    uploaded_scan = st.file_uploader("通帳スキャン画像をアップロードしてください", type=["jpg", "jpeg", "png"], key="scan_upload")
    if uploaded_scan is not None:
        st.image(uploaded_scan, caption="アップロード画像", use_container_width=True)
        if st.button("AIで通帳を読み取る", type="primary"):
            try:
                scan_rows = scan_passbook_to_json(uploaded_scan.getvalue(), uploaded_scan.name, model_name=scan_model)
                st.session_state["scan_json_store"] = scan_rows
                bank_df_scan = normalize_passbook_json(scan_rows, year_prefix=scan_year_prefix)
                st.session_state["bank_df_store"] = bank_df_scan
                st.success(f"AI読取が完了しました。{len(bank_df_scan)} 件の明細を正規化しました。")
            except Exception as e:
                st.error(f"通帳スキャンエラー: {e}")
        if st.session_state["scan_json_store"] is not None:
            with st.expander("AI読取JSONを確認"):
                st.json(st.session_state["scan_json_store"])
            if st.session_state["bank_df_store"] is not None and not st.session_state["bank_df_store"].empty:
                with st.expander("正規化後データを確認"):
                    preview_scan = st.session_state["bank_df_store"].copy()
                    preview_scan["date"] = preview_scan["date"].dt.strftime("%Y-%m-%d")
                    preview_scan["amount"] = preview_scan["amount"].map(fmt)
                    if "balance" in preview_scan.columns:
                        preview_scan["balance"] = preview_scan["balance"].map(fmt)
                    st.dataframe(preview_scan, use_container_width=True)

if font_msg:
    st.info(font_msg)
if st.session_state["bank_df_store"] is not None:
    bank_df = st.session_state["bank_df_store"]

mode = st.radio("モード選択", ["通常モード", "案件モード"], horizontal=True)

events_df = build_event_df(st.session_state["events_store"])

project_initial_balance = 0

if mode == "通常モード":
    st.subheader("2. 今後の予定を出来事として追加")
else:
    st.subheader("2. 案件の入金・支払予定を追加")

if mode == "案件モード":
    project_initial_balance_man = st.number_input("初期残高（万円）", min_value=0, value=300, step=10)
    project_initial_balance = project_initial_balance_man * 10000

excel_tab1, excel_tab2 = st.tabs(["手入力", "Excel一括取込"])

with excel_tab1:
    col1, col2, col3 = st.columns([1.2, 1, 2.8])

    with col1:
        ev_date = st.date_input("日付", value=date.today(), key="new_date")

    with col2:
        ev_kind = st.selectbox("種別", ["入金", "出金"], key="new_kind")
        ev_repeat = st.selectbox("繰返し", ["単発", "毎月"], key="new_repeat")

    with col3:
        st.markdown("**金額入力**")
        unit = st.radio("単位", ["円", "百円", "千円", "万円"], horizontal=True, key="new_unit")
        unit_map = {"円": 1, "百円": 100, "千円": 1000, "万円": 10000}
        amount_digits = st.session_state["amount_digits"] or "0"
        st.text_input("入力中の数字", value=amount_digits, disabled=True)

        keypad_rows = [["7", "8", "9", "000"], ["4", "5", "6", "00"], ["1", "2", "3", "0"]]
        for row_idx, row in enumerate(keypad_rows):
            cols = st.columns(4)
            for i, key in enumerate(row):
                if cols[i].button(key, key=f"key_{row_idx}_{key}"):
                    append_digits(key)

        cols = st.columns(4)
        if cols[0].button("C", key="clear_key"):
            clear_digits()
        if cols[1].button("⌫", key="back_key"):
            backspace_digits()
        if cols[2].button("+千円", key="plus_1000"):
            st.session_state["amount_digits"] = str(int(st.session_state["amount_digits"] or "0") + 1000)
        if cols[3].button("+万円", key="plus_10000"):
            st.session_state["amount_digits"] = str(int(st.session_state["amount_digits"] or "0") + 10000)

        preview_amount = int(st.session_state["amount_digits"] or "0") * unit_map[unit]
        st.markdown(f"**確定金額:** {fmt(preview_amount)} 円")

    memo_col1, memo_col2 = st.columns([1, 1.6])

    with memo_col1:
        memo_preset = st.selectbox(
            "メモ候補",
            ["", "売上増加", "設備投資", "賞与", "税金", "借入返済", "入金遅延", "その他"],
            key="new_memo_preset"
        )

    with memo_col2:
        memo_free = st.text_input(
            "自由メモ",
            value="",
            placeholder="例：6月に機械修理代を支払う",
            key="new_memo_free"
        )

    memo_final = memo_free.strip() if memo_free.strip() else memo_preset

    button_col1, button_col2 = st.columns([1, 3])

    with button_col1:
        if st.button("予定を追加", type="primary"):
            if preview_amount <= 0:
                st.warning("金額を入力してください。")
            else:
                new_id = st.session_state["next_event_id"]
                new_row = pd.DataFrame([{
                    "id": new_id,
                    "date": pd.Timestamp(ev_date),
                    "kind": ev_kind,
                    "amount": preview_amount,
                    "repeat": ev_repeat,
                    "memo": memo_final
                }])
                st.session_state["events_store"] = pd.concat(
                    [st.session_state["events_store"], new_row],
                    ignore_index=True
                )
                st.session_state["next_event_id"] += 1
                clear_digits()
                st.success("予定を追加しました。")
                st.rerun()

    with button_col2:
        st.caption(event_summary_text(ev_kind, preview_amount, ev_repeat, memo_final))

with excel_tab2:
    st.download_button(
        "予定Excelテンプレをダウンロード",
        data=create_schedule_template_bytes(),
        file_name="資金繰り予定テンプレ_改善版.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.markdown(
        '<div class="small-note">必要列は「日付 / 種別 / 金額 / 繰返し / メモ」です。</div>',
        unsafe_allow_html=True
    )

    uploaded_schedule = st.file_uploader(
        "予定Excelをアップロードしてください",
        type=["xlsx"],
        key="schedule_excel_upload"
    )

    if uploaded_schedule is not None:
        try:
            schedule_df = load_schedule_excel(uploaded_schedule)
            preview_schedule = schedule_df.copy()
            preview_schedule["date"] = preview_schedule["date"].dt.strftime("%Y-%m-%d")
            preview_schedule["amount"] = preview_schedule["amount"].map(fmt)

            st.dataframe(
                preview_schedule.rename(columns={
                    "date": "日付",
                    "kind": "種別",
                    "amount": "金額（円）",
                    "repeat": "繰返し",
                    "memo": "メモ"
                }),
                use_container_width=True
            )

            if st.button("このExcel予定を追加"):
                add_df = schedule_df.copy()
                add_df["id"] = range(
                    st.session_state["next_event_id"],
                    st.session_state["next_event_id"] + len(add_df)
                )
                add_df = add_df[["id", "date", "kind", "amount", "repeat", "memo"]]
                st.session_state["events_store"] = pd.concat(
                    [st.session_state["events_store"], add_df],
                    ignore_index=True
                )
                st.session_state["next_event_id"] += len(add_df)
                st.success(f"{len(add_df)} 件の予定を追加しました。")
                st.rerun()

        except Exception as e:
            st.error(f"予定Excel取込エラー: {e}")

events_df = build_event_df(st.session_state["events_store"])

st.subheader("2-2. 追加した予定の確認・編集・削除")
if not events_df.empty:
    display_events = events_df.copy()
    display_events["date"] = display_events["date"].dt.strftime("%Y-%m-%d")
    display_events["amount"] = display_events["amount"].map(fmt)

    st.dataframe(
        display_events.rename(columns={
            "id": "No.",
            "date": "日付",
            "kind": "種別",
            "amount": "金額（円）",
            "repeat": "繰返し",
            "memo": "メモ",
        }),
        use_container_width=True,
    )

    event_labels = [format_event_label(r) for _, r in events_df.iterrows()]
    selected_index = st.selectbox(
        "対象予定を選択",
        range(len(event_labels)),
        format_func=lambda i: event_labels[i],
    )

    a1, a2, a3 = st.columns([1, 1, 3])

    with a1:
        if st.button("この予定を編集"):
            st.session_state["edit_target_id"] = selected_index
            st.rerun()

    with a2:
        if st.button("この予定を削除"):
            updated = build_event_df(st.session_state["events_store"]).reset_index(drop=True)
            updated = updated.drop(index=selected_index).reset_index(drop=True)
            st.session_state["events_store"] = updated

            if st.session_state["edit_target_id"] == selected_index:
                st.session_state["edit_target_id"] = None

            st.success("選択した予定を削除しました。")
            st.rerun()

    with a3:
        if st.button("予定をすべて削除"):
            st.session_state["events_store"] = pd.DataFrame(
                columns=["id", "date", "kind", "amount", "repeat", "memo"]
            )
            st.session_state["edit_target_id"] = None
            st.success("予定をすべて削除しました。")
            st.rerun()

    if st.session_state["edit_target_id"] is not None:
        edit_index = st.session_state["edit_target_id"]

        if 0 <= edit_index < len(events_df):
            edit_row = events_df.iloc[edit_index]

            st.markdown(f"#### No.{int(edit_row['id'])} の編集")

            e1, e2, e3 = st.columns([1.2, 1, 2.2])

            with e1:
                edit_date = st.date_input(
                    "編集日付",
                    value=pd.Timestamp(edit_row["date"]).date(),
                    key=f"edit_date_{edit_index}",
                )

            with e2:
                edit_kind = st.selectbox(
                    "編集種別",
                    ["入金", "出金"],
                    index=0 if edit_row["kind"] == "入金" else 1,
                    key=f"edit_kind_{edit_index}",
                )
                edit_repeat = st.selectbox(
                    "編集繰返し",
                    ["単発", "毎月"],
                    index=0 if edit_row["repeat"] == "単発" else 1,
                    key=f"edit_repeat_{edit_index}",
                )

            with e3:
                edit_amount = st.number_input(
                    "編集金額（円）",
                    value=int(edit_row["amount"]),
                    step=10000,
                    key=f"edit_amount_{edit_index}",
                )

            m1, m2 = st.columns([1, 1.6])

            with m1:
                if edit_kind == "入金":
                    preset_list = ["", "売上"]
                else:
                    preset_list = [
                        "",
                        "仕入支払",
                        "外注費",
                        "人件費",
                        "家賃",
                        "設備投資",
                        "借入返済",
                        "利息支払",
                        "税金",
                        "賞与",
                        "修繕費",
                    ]

                current_memo = str(edit_row["memo"]) if pd.notna(edit_row["memo"]) else ""
                preset_index = preset_list.index(current_memo) if current_memo in preset_list else 0

                edit_memo_preset = st.selectbox(
                    "編集メモ候補",
                    preset_list,
                    index=preset_index,
                    key=f"edit_memo_preset_{edit_index}",
                )

            with m2:
                current_memo = str(edit_row["memo"]) if pd.notna(edit_row["memo"]) else ""
                edit_memo_free = st.text_input(
                    "編集自由メモ",
                    value="" if current_memo in preset_list else current_memo,
                    placeholder="必要に応じて入力",
                    key=f"edit_memo_free_{edit_index}",
                )

            edit_memo_final = edit_memo_free.strip() if edit_memo_free.strip() else edit_memo_preset

            b1, b2 = st.columns([1, 3])

            with b1:
                if st.button("この内容で更新", type="primary", key=f"update_event_{edit_index}"):
                    updated = build_event_df(st.session_state["events_store"]).reset_index(drop=True)

                    updated.loc[edit_index, "date"] = pd.Timestamp(edit_date)
                    updated.loc[edit_index, "kind"] = edit_kind
                    updated.loc[edit_index, "amount"] = int(edit_amount)
                    updated.loc[edit_index, "repeat"] = edit_repeat
                    updated.loc[edit_index, "memo"] = edit_memo_final

                    st.session_state["events_store"] = updated
                    st.session_state["edit_target_id"] = None
                    st.success("予定を更新しました。")
                    st.rerun()

            with b2:
                if st.button("編集をキャンセル", key=f"cancel_event_{edit_index}"):
                    st.session_state["edit_target_id"] = None
                    st.rerun()
else:
    st.info("まだ予定は追加されていません。")

bank_df = st.session_state["bank_df_store"] if st.session_state["bank_df_store"] is not None else None
events_df = build_event_df(st.session_state["events_store"])

st.subheader("3. 再生成結果")

forecast = None
current_balance = 0
event_detail = pd.DataFrame()

if mode == "通常モード":
    if bank_df is not None and not bank_df.empty:
        forecast, current_balance, event_detail = build_forecast(
            bank_df,
            events_df,
            months_back=months_back,
            horizon_months=horizon_months,
            safety_threshold=safety_threshold
        )
else:
    forecast, current_balance, event_detail = build_project_forecast_from_events(
        initial_balance=project_initial_balance,
        events_df=events_df,
        horizon_months=horizon_months,
        safety_threshold=safety_threshold
    )

if forecast is not None and not forecast.empty:
    danger_info_month, danger_analysis = analyze_danger_reason(forecast, event_detail)
    danger_df = forecast[forecast["danger"]]

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(f'<div class="kpi-card"><div class="kpi-title">現在残高</div><div class="kpi-value">{fmt_man(current_balance)} 万円</div></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="kpi-card"><div class="kpi-title">12か月最低残高</div><div class="kpi-value">{fmt_man(forecast["ending_balance"].min())} 万円</div></div>', unsafe_allow_html=True)
    with c3:
        st.markdown(f'<div class="kpi-card"><div class="kpi-title">危険月数</div><div class="kpi-value">{int(forecast["danger"].sum())}</div></div>', unsafe_allow_html=True)

    if danger_info_month is not None:
        st.error(f"⚠ {danger_info_month['month']:%Y-%m} に安全残高 {fmt_man(safety_threshold)} 万円を下回ります。月末残高は {fmt_man(danger_info_month['ending_balance'])} 万円です。")

    max_for_scale = max(abs(forecast["ending_balance"]).max(), abs(safety_threshold), 1)
    scale, unit = get_scale_info(max_for_scale)
    forecast["scaled_balance"] = forecast["ending_balance"] / scale
    scaled_threshold = safety_threshold / scale

    fig, ax = plt.subplots(figsize=(12, 4.8))
    ax.plot(forecast["month"], forecast["scaled_balance"], marker="o", linewidth=3, markersize=7)
    ax.axhline(y=scaled_threshold, linestyle="--", linewidth=2)
    ax.set_title("向こう12か月の月末残高推移", fontsize=14, fontweight="bold")
    ax.set_xlabel("月")
    ax.set_ylabel(f"残高（{unit}）")
    ax.grid(True, alpha=0.35)
    plt.xticks(rotation=45)
    st.pyplot(fig)

    table_df = pd.DataFrame({"項目": ["前月繰越", "入金合計", "出金合計", "収支", "月末残高"]})
    for _, r in forecast.iterrows():
        col = r["month"].strftime("%Y-%m")
        table_df[col] = [r["carry_in"], r["inflow_total"], r["outflow_total"], r["net"], r["ending_balance"]]

    styled_table = (
        table_df.set_index("項目")
        .style
        .format(fmt)
        .apply(highlight_row, axis=1)
        .map(highlight_cashflow, subset=pd.IndexSlice[["収支"], :])
    )

    st.markdown("#### 向こう12か月分の資金繰り表")
    st.dataframe(styled_table, use_container_width=True, height=280)

    if not danger_df.empty:
        danger_out = danger_df[["month", "ending_balance"]].copy()
        danger_out["month"] = danger_out["month"].dt.strftime("%Y-%m")
        st.markdown("#### 危険月")
        st.dataframe(
            danger_out.rename(columns={"month": "月", "ending_balance": "月末残高（円）"}).style.format({"月末残高（円）": fmt}),
            use_container_width=True
        )

    if danger_analysis is not None:
        st.markdown("#### 危険月の主な要因")
        st.text(danger_analysis["summary"])
        if danger_analysis["table"] is not None:
            st.dataframe(danger_analysis["table"], use_container_width=True)
    else:
        st.success("安全残高を下回る月はありません。")

    output = io.BytesIO()

    forecast_export = forecast.copy()
    forecast_export = forecast_export.drop(columns=["scaled_balance"], errors="ignore")
    forecast_export["month"] = pd.to_datetime(forecast_export["month"]).dt.strftime("%Y-%m")

    forecast_export = forecast_export.rename(columns={
        "month": "年月",
        "base_inflow": "基本入金",
        "base_outflow": "基本出金",
        "event_inflow": "イベント入金",
        "event_outflow": "イベント出金",
        "inflow_total": "総入金",
        "outflow_total": "総出金",
        "net": "収支",
        "carry_in": "期首残高",
        "ending_balance": "期末残高",
        "danger": "危険判定"
    })

    forecast_export["危険判定"] = forecast_export["危険判定"].map({
        True: "危険",
        False: "正常"
    })

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        forecast_export.to_excel(writer, index=False, sheet_name="資金繰り予測")
        ws = writer.book["資金繰り予測"]

        widths = {
            "A": 12, "B": 15, "C": 15, "D": 15, "E": 15,
            "F": 15, "G": 15, "H": 12, "I": 15, "J": 15, "K": 10
        }
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column_letter != "K" and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'

        table_df.to_excel(writer, index=False, sheet_name="月次一覧")
        ws2 = writer.book["月次一覧"]
        ws2.column_dimensions["A"].width = 12
        for col_idx in range(2, ws2.max_column + 1):
            col_letter = ws2.cell(row=1, column=col_idx).column_letter
            ws2.column_dimensions[col_letter].width = 14
        for row in ws2.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'

        if not events_df.empty:
            events_df.to_excel(writer, index=False, sheet_name="予定イベント")
            ws3 = writer.book["予定イベント"]
            widths = {"A": 8, "B": 12, "C": 10, "D": 12, "E": 10, "F": 24}
            for col, w in widths.items():
                ws3.column_dimensions[col].width = w
            for row in ws3.iter_rows(min_row=2):
                for cell in row:
                    if cell.column_letter == "B":
                        cell.number_format = "yyyy-mm-dd"
                    elif cell.column_letter == "D" and isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0'

        if not event_detail.empty:
            event_detail.to_excel(writer, index=False, sheet_name="イベント詳細")
            ws4 = writer.book["イベント詳細"]
            widths = {"A": 12, "B": 8, "C": 10, "D": 12, "E": 10, "F": 24}
            for col, w in widths.items():
                ws4.column_dimensions[col].width = w
            for row in ws4.iter_rows(min_row=2):
                for cell in row:
                    if cell.column_letter == "D" and isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0'

        if mode == "通常モード" and bank_df is not None and not bank_df.empty:
            bank_df.to_excel(writer, index=False, sheet_name="実績データ")
            ws5 = writer.book["実績データ"]
            default_widths = {"A": 12, "B": 12, "C": 14, "D": 20, "E": 24}
            for col, w in default_widths.items():
                ws5.column_dimensions[col].width = w
            for row in ws5.iter_rows(min_row=2):
                for cell in row:
                    if cell.column_letter == "A":
                        cell.number_format = "yyyy-mm-dd"
                    elif cell.column_letter in ["B", "C"] and isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0'

    from datetime import datetime

    st.download_button(
        "結果をExcelでダウンロード",
        data=output.getvalue(),
        file_name=f"資金繰り_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

else:
    st.info("通常モードでは実績データを取り込むと結果を表示します。")
